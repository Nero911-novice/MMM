# export_manager.py
"""
–°–∏—Å—Ç–µ–º–∞ —ç–∫—Å–ø–æ—Ä—Ç–∞ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ Marketing Mix Model –≤ Excel –∏ PDF.
–í–µ—Ä—Å–∏—è 2.1 —Å —Ä–∞—Å—à–∏—Ä–µ–Ω–Ω—ã–º–∏ –≤–æ–∑–º–æ–∂–Ω–æ—Å—Ç—è–º–∏ –æ—Ç—á–µ—Ç–Ω–æ—Å—Ç–∏ –∏ –∏—Å–ø—Ä–∞–≤–ª–µ–Ω–Ω—ã–º–∏ –æ—à–∏–±–∫–∞–º–∏.
"""

import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO
import base64

# Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF export
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from config import COLOR_PALETTE, format_number

class ExportManager:
    """–ö–ª–∞—Å—Å –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ MMM –≤ —Ä–∞–∑–ª–∏—á–Ω—ã–µ —Ñ–æ—Ä–º–∞—Ç—ã."""
    
    def __init__(self):
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
    def check_dependencies(self):
        """–ü—Ä–æ–≤–µ—Ä–∫–∞ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –±–∏–±–ª–∏–æ—Ç–µ–∫ –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞."""
        return {
            'excel': EXCEL_AVAILABLE,
            'pdf': PDF_AVAILABLE
        }
    
    def _safe_get(self, data, key, default=None):
        """–ë–µ–∑–æ–ø–∞—Å–Ω–æ–µ –ø–æ–ª—É—á–µ–Ω–∏–µ –∑–Ω–∞—á–µ–Ω–∏—è –∏–∑ —Å–ª–æ–≤–∞—Ä—è."""
        if data is None:
            return default
        if isinstance(data, dict):
            return data.get(key, default)
        return default
    
    def _safe_access_df(self, data, operation='len'):
        """–ë–µ–∑–æ–ø–∞—Å–Ω—ã–π –¥–æ—Å—Ç—É–ø –∫ DataFrame."""
        if data is None:
            return 0 if operation == 'len' else False
        if isinstance(data, pd.DataFrame):
            if operation == 'len':
                return len(data)
            elif operation == 'empty':
                return data.empty
            elif operation == 'exists':
                return True
        return 0 if operation == 'len' else False
    
    def export_to_excel(self, model_results, filename=None):
        """
        –≠–∫—Å–ø–æ—Ä—Ç —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ MMM –≤ Excel —Ñ–∞–π–ª.
        
        Parameters:
        -----------
        model_results : dict
            –°–ª–æ–≤–∞—Ä—å —Å —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏ –º–æ–¥–µ–ª–∏ (contributions, roas_data, metrics, etc.)
        filename : str
            –ò–º—è —Ñ–∞–π–ª–∞ (–æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–æ)
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("–î–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞ –≤ Excel —É—Å—Ç–∞–Ω–æ–≤–∏—Ç–µ: pip install openpyxl xlsxwriter")
        
        if filename is None:
            filename = f"MMM_Report_{self.timestamp}.xlsx"
        
        # –ü—Ä–æ–≤–µ—Ä–∫–∞ –≤—Ö–æ–¥–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö
        if model_results is None:
            model_results = {}
        
        # –°–æ–∑–¥–∞–Ω–∏–µ –±—É—Ñ–µ—Ä–∞ –≤ –ø–∞–º—è—Ç–∏
        buffer = BytesIO()
        
        # –°–æ–∑–¥–∞–Ω–∏–µ Excel —Ñ–∞–π–ª–∞
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            
            # –õ–∏—Å—Ç 1: –°–≤–æ–¥–∫–∞
            self._create_summary_sheet(writer, model_results)
            
            # –õ–∏—Å—Ç 2: –î–µ–∫–æ–º–ø–æ–∑–∏—Ü–∏—è –ø—Ä–æ–¥–∞–∂
            contributions = self._safe_get(model_results, 'contributions', {})
            if contributions:
                self._create_decomposition_sheet(writer, contributions)
            
            # –õ–∏—Å—Ç 3: ROAS –∞–Ω–∞–ª–∏–∑
            roas_data = self._safe_get(model_results, 'roas_data')
            if roas_data is not None and not self._safe_access_df(roas_data, 'empty'):
                self._create_roas_sheet(writer, roas_data)
            
            # –õ–∏—Å—Ç 4: –ú–µ—Ç—Ä–∏–∫–∏ –∫–∞—á–µ—Å—Ç–≤–∞ –º–æ–¥–µ–ª–∏
            model_metrics = self._safe_get(model_results, 'model_metrics', {})
            if model_metrics:
                self._create_metrics_sheet(writer, model_metrics)
            
            # –õ–∏—Å—Ç 5: –û–ø—Ç–∏–º–∏–∑–∞—Ü–∏—è –±—é–¥–∂–µ—Ç–∞
            optimization_results = self._safe_get(model_results, 'optimization_results')
            if optimization_results:
                self._create_optimization_sheet(writer, optimization_results)
            
            # –õ–∏—Å—Ç 6: –°—Ü–µ–Ω–∞—Ä–∏–∏
            scenarios = self._safe_get(model_results, 'scenarios')
            if scenarios:
                self._create_scenarios_sheet(writer, scenarios)
            
            # –õ–∏—Å—Ç 7: –î–∞–Ω–Ω—ã–µ –¥–ª—è –∞–Ω–∞–ª–∏–∑–∞
            raw_data = self._safe_get(model_results, 'raw_data')
            if raw_data is not None and not self._safe_access_df(raw_data, 'empty'):
                self._create_data_sheet(writer, raw_data)
            
            # –õ–∏—Å—Ç 8: –ò–Ω—Å–∞–π—Ç—ã –∏ —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏
            self._create_insights_sheet(writer, model_results)
        
        buffer.seek(0)
        return buffer.getvalue(), filename
    
    def _create_summary_sheet(self, writer, model_results):
        """–°–æ–∑–¥–∞–Ω–∏–µ —Å–≤–æ–¥–Ω–æ–≥–æ –ª–∏—Å—Ç–∞."""
        ws_name = "–°–≤–æ–¥–∫–∞"
        
        # –°–æ–∑–¥–∞–Ω–∏–µ —Å–≤–æ–¥–Ω–æ–π –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏ —Å –±–µ–∑–æ–ø–∞—Å–Ω—ã–º –¥–æ—Å—Ç—É–ø–æ–º
        summary_data = {
            '–ü–∞—Ä–∞–º–µ—Ç—Ä': [
                '–î–∞—Ç–∞ —Å–æ–∑–¥–∞–Ω–∏—è –æ—Ç—á–µ—Ç–∞',
                '–ö–∞—á–µ—Å—Ç–≤–æ –º–æ–¥–µ–ª–∏ (R¬≤)',
                '–¢–æ—á–Ω–æ—Å—Ç—å –ø—Ä–æ–≥–Ω–æ–∑–∞ (%)',
                '–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –º–µ–¥–∏–∞-–∫–∞–Ω–∞–ª–æ–≤',
                '–ü–µ—Ä–∏–æ–¥ –∞–Ω–∞–ª–∏–∑–∞',
                '–û–±—â–∏–π –±—é–¥–∂–µ—Ç (—Ç—ã—Å. —Ä—É–±)',
                '–°—Ä–µ–¥–Ω–∏–π ROAS',
                '–ë–∞–∑–æ–≤–∞—è –ª–∏–Ω–∏—è (%)',
                '–°—Ç–∞—Ç—É—Å –º–æ–¥–µ–ª–∏'
            ],
            '–ó–Ω–∞—á–µ–Ω–∏–µ': [
                datetime.now().strftime("%d.%m.%Y %H:%M"),
                f"{self._safe_get(model_results, 'r2_score', 0):.3f}",
                f"{self._safe_get(model_results, 'accuracy', 0):.1f}%",
                len(self._safe_get(model_results, 'media_channels', [])),
                self._safe_get(model_results, 'analysis_period', '–ù/–î'),
                f"{self._safe_get(model_results, 'total_budget', 0)/1000:.0f}",
                f"{self._safe_get(model_results, 'avg_roas', 0):.2f}",
                f"{self._safe_get(model_results, 'base_contribution_pct', 0):.1f}%",
                self._safe_get(model_results, 'model_status', '–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–æ')
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name=ws_name, index=False, startrow=2)
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
        workbook = writer.book
        worksheet = writer.sheets[ws_name]
        
        # –ó–∞–≥–æ–ª–æ–≤–æ–∫
        worksheet['A1'] = '–û–¢–ß–ï–¢ –ü–û MARKETING MIX MODEL'
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet.merge_cells('A1:B1')
        
        # –°—Ç–∏–ª–∏–∑–∞—Ü–∏—è —Ç–∞–±–ª–∏—Ü—ã
        self._style_excel_table(worksheet, start_row=3, end_row=3+len(summary_df), 
                               start_col=1, end_col=2)
    
    def _create_decomposition_sheet(self, writer, contributions):
        """–°–æ–∑–¥–∞–Ω–∏–µ –ª–∏—Å—Ç–∞ –¥–µ–∫–æ–º–ø–æ–∑–∏—Ü–∏–∏ –ø—Ä–æ–¥–∞–∂."""
        ws_name = "–î–µ–∫–æ–º–ø–æ–∑–∏—Ü–∏—è"
        
        # –ü–æ–¥–≥–æ—Ç–æ–≤–∫–∞ –¥–∞–Ω–Ω—ã—Ö
        contrib_data = []
        total_contribution = sum(contributions.values()) if contributions else 1
        
        for channel, value in contributions.items():
            contrib_data.append({
                '–ö–∞–Ω–∞–ª': str(channel).replace('_spend', '').replace('_', ' ').title(),
                '–í–∫–ª–∞–¥ (—à—Ç)': int(float(value)) if value is not None else 0,
                '–î–æ–ª—è (%)': round(float(value) / total_contribution * 100, 1) if total_contribution > 0 and value is not None else 0,
                '–¢–∏–ø': '–ë–∞–∑–æ–≤–∞—è –ª–∏–Ω–∏—è' if channel == 'Base' else '–ú–µ–¥–∏–∞-–∫–∞–Ω–∞–ª'
            })
        
        contrib_df = pd.DataFrame(contrib_data)
        contrib_df = contrib_df.sort_values('–í–∫–ª–∞–¥ (—à—Ç)', ascending=False)
        
        # –ó–∞–ø–∏—Å—å –≤ Excel
        contrib_df.to_excel(writer, sheet_name=ws_name, index=False, startrow=2)
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
        worksheet = writer.sheets[ws_name]
        worksheet['A1'] = '–î–ï–ö–û–ú–ü–û–ó–ò–¶–ò–Ø –ü–†–û–î–ê–ñ –ü–û –ö–ê–ù–ê–õ–ê–ú'
        worksheet['A1'].font = Font(size=14, bold=True)
        worksheet.merge_cells('A1:D1')
        
        self._style_excel_table(worksheet, start_row=3, end_row=3+len(contrib_df), 
                               start_col=1, end_col=4)
        
        # –î–æ–±–∞–≤–ª–µ–Ω–∏–µ –¥–∏–∞–≥—Ä–∞–º–º—ã
        self._add_chart_to_sheet(worksheet, contrib_df, 'F3', chart_type='bar')
    
    def _create_roas_sheet(self, writer, roas_data):
        """–°–æ–∑–¥–∞–Ω–∏–µ –ª–∏—Å—Ç–∞ ROAS –∞–Ω–∞–ª–∏–∑–∞."""
        ws_name = "ROAS"
        
        if isinstance(roas_data, pd.DataFrame) and not roas_data.empty:
            # –î–æ–±–∞–≤–ª–µ–Ω–∏–µ –∏–Ω—Ç–µ—Ä–ø—Ä–µ—Ç–∞—Ü–∏–∏ ROAS
            roas_with_interpretation = roas_data.copy()
            roas_with_interpretation['–û—Ü–µ–Ω–∫–∞'] = roas_with_interpretation['ROAS'].apply(
                lambda x: '–û—Ç–ª–∏—á–Ω–æ' if x >= 3.0 else '–•–æ—Ä–æ—à–æ' if x >= 2.0 else '–ü—Ä–∏–µ–º–ª–µ–º–æ' if x >= 1.5 else '–ü–ª–æ—Ö–æ'
            )
            roas_with_interpretation['–†–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏—è'] = roas_with_interpretation['ROAS'].apply(
                lambda x: '–£–≤–µ–ª–∏—á–∏—Ç—å –±—é–¥–∂–µ—Ç' if x >= 3.0 else '–°–æ—Ö—Ä–∞–Ω–∏—Ç—å —É—Ä–æ–≤–µ–Ω—å' if x >= 2.0 else '–û–ø—Ç–∏–º–∏–∑–∏—Ä–æ–≤–∞—Ç—å' if x >= 1.5 else '–°–æ–∫—Ä–∞—Ç–∏—Ç—å/–ø–µ—Ä–µ–Ω–∞—Å—Ç—Ä–æ–∏—Ç—å'
            )
            
            roas_with_interpretation.to_excel(writer, sheet_name=ws_name, index=False, startrow=2)
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
            worksheet = writer.sheets[ws_name]
            worksheet['A1'] = '–ê–ù–ê–õ–ò–ó ROAS –ü–û –ö–ê–ù–ê–õ–ê–ú'
            worksheet['A1'].font = Font(size=14, bold=True)
            worksheet.merge_cells('A1:F1')
            
            self._style_excel_table(worksheet, start_row=3, end_row=3+len(roas_with_interpretation), 
                                   start_col=1, end_col=6)
    
    def _create_metrics_sheet(self, writer, model_metrics):
        """–°–æ–∑–¥–∞–Ω–∏–µ –ª–∏—Å—Ç–∞ –º–µ—Ç—Ä–∏–∫ –∫–∞—á–µ—Å—Ç–≤–∞ –º–æ–¥–µ–ª–∏."""
        ws_name = "–ö–∞—á–µ—Å—Ç–≤–æ –º–æ–¥–µ–ª–∏"
        
        metrics_data = []
        for metric, value in model_metrics.items():
            if isinstance(value, (int, float)):
                if '–¢–æ—á–Ω–æ—Å—Ç—å' in metric or '%' in metric:
                    formatted_value = f"{value:.1f}%"
                elif '–ö–∞—á–µ—Å—Ç–≤–æ' in metric:
                    formatted_value = f"{value:.3f}"
                else:
                    formatted_value = f"{value:,.0f}"
            else:
                formatted_value = str(value)
                
            metrics_data.append({
                '–ú–µ—Ç—Ä–∏–∫–∞': metric,
                '–ó–Ω–∞—á–µ–Ω–∏–µ': formatted_value,
                '–ò–Ω—Ç–µ—Ä–ø—Ä–µ—Ç–∞—Ü–∏—è': self._interpret_metric(metric, value)
            })
        
        metrics_df = pd.DataFrame(metrics_data)
        metrics_df.to_excel(writer, sheet_name=ws_name, index=False, startrow=2)
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
        worksheet = writer.sheets[ws_name]
        worksheet['A1'] = '–ú–ï–¢–†–ò–ö–ò –ö–ê–ß–ï–°–¢–í–ê –ú–û–î–ï–õ–ò'
        worksheet['A1'].font = Font(size=14, bold=True)
        worksheet.merge_cells('A1:C1')
        
        self._style_excel_table(worksheet, start_row=3, end_row=3+len(metrics_df), 
                               start_col=1, end_col=3)
    
    def _create_optimization_sheet(self, writer, optimization_results):
        """–°–æ–∑–¥–∞–Ω–∏–µ –ª–∏—Å—Ç–∞ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏."""
        ws_name = "–û–ø—Ç–∏–º–∏–∑–∞—Ü–∏—è"
        
        allocation = self._safe_get(optimization_results, 'allocation', {})
        if allocation:
            opt_data = []
            total_budget = sum(allocation.values())
            
            for channel, budget in allocation.items():
                opt_data.append({
                    '–ö–∞–Ω–∞–ª': str(channel).replace('_spend', '').replace('_', ' ').title(),
                    '–û–ø—Ç–∏–º–∞–ª—å–Ω—ã–π –±—é–¥–∂–µ—Ç': int(float(budget)) if budget is not None else 0,
                    '–î–æ–ª—è (%)': round(float(budget) / total_budget * 100, 1) if total_budget > 0 and budget is not None else 0
                })
            
            opt_df = pd.DataFrame(opt_data)
            opt_df.to_excel(writer, sheet_name=ws_name, index=False, startrow=2)
            
            # –î–æ–±–∞–≤–ª–µ–Ω–∏–µ —Å–≤–æ–¥–∫–∏ –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏
            worksheet = writer.sheets[ws_name]
            summary_start_row = len(opt_df) + 5
            summary_data = [
                ['–ü—Ä–æ–≥–Ω–æ–∑–∏—Ä—É–µ–º—ã–µ –ø—Ä–æ–¥–∞–∂–∏', f"{self._safe_get(optimization_results, 'predicted_sales', 0):,.0f}"],
                ['–ü—Ä–æ–≥–Ω–æ–∑–∏—Ä—É–µ–º—ã–π ROAS', f"{self._safe_get(optimization_results, 'predicted_roas', 0):.2f}"],
                ['–ü—Ä–æ–≥–Ω–æ–∑–∏—Ä—É–µ–º—ã–π ROI', f"{self._safe_get(optimization_results, 'predicted_roi', 0):.2f}"],
                ['–û–±—â–∏–π –±—é–¥–∂–µ—Ç', f"{self._safe_get(optimization_results, 'total_budget_used', 0):,.0f}"],
                ['–ú–µ—Ç–æ–¥ –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏', self._safe_get(optimization_results, 'optimization_method', '–ù/–î')]
            ]
            
            for i, (param, value) in enumerate(summary_data):
                worksheet.cell(row=summary_start_row + i, column=1, value=param)
                worksheet.cell(row=summary_start_row + i, column=2, value=value)
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
            worksheet['A1'] = '–†–ï–ó–£–õ–¨–¢–ê–¢–´ –û–ü–¢–ò–ú–ò–ó–ê–¶–ò–ò –ë–Æ–î–ñ–ï–¢–ê'
            worksheet['A1'].font = Font(size=14, bold=True)
            worksheet.merge_cells('A1:C1')
    
    def _create_scenarios_sheet(self, writer, scenarios):
        """–°–æ–∑–¥–∞–Ω–∏–µ –ª–∏—Å—Ç–∞ —Å—Ü–µ–Ω–∞—Ä–Ω–æ–≥–æ –∞–Ω–∞–ª–∏–∑–∞."""
        ws_name = "–°—Ü–µ–Ω–∞—Ä–∏–∏"
        
        scenarios_df = pd.DataFrame(scenarios).T
        scenarios_df.to_excel(writer, sheet_name=ws_name, startrow=2)
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
        worksheet = writer.sheets[ws_name]
        worksheet['A1'] = '–°–¶–ï–ù–ê–†–ù–´–ô –ê–ù–ê–õ–ò–ó'
        worksheet['A1'].font = Font(size=14, bold=True)
        worksheet.merge_cells('A1:D1')
    
    def _create_data_sheet(self, writer, raw_data):
        """–°–æ–∑–¥–∞–Ω–∏–µ –ª–∏—Å—Ç–∞ —Å –∏—Å—Ö–æ–¥–Ω—ã–º–∏ –¥–∞–Ω–Ω—ã–º–∏."""
        ws_name = "–î–∞–Ω–Ω—ã–µ"
        
        if isinstance(raw_data, pd.DataFrame):
            raw_data.to_excel(writer, sheet_name=ws_name, index=False)
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
            worksheet = writer.sheets[ws_name]
            for col in range(1, len(raw_data.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def _create_insights_sheet(self, writer, model_results):
        """–°–æ–∑–¥–∞–Ω–∏–µ –ª–∏—Å—Ç–∞ —Å –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏–º–∏ –∏–Ω—Å–∞–π—Ç–∞–º–∏."""
        ws_name = "–ò–Ω—Å–∞–π—Ç—ã"
        
        # –ì–µ–Ω–µ—Ä–∞—Ü–∏—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏—Ö –∏–Ω—Å–∞–π—Ç–æ–≤
        insights = self.create_automated_insights(model_results)
        
        insights_data = []
        
        # –ò–Ω—Å–∞–π—Ç—ã –ø–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å–Ω–æ—Å—Ç–∏
        for insight in insights['performance_insights']:
            insights_data.append({'–ö–∞—Ç–µ–≥–æ—Ä–∏—è': '–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å', '–ò–Ω—Å–∞–π—Ç': insight})
        
        # –í–æ–∑–º–æ–∂–Ω–æ—Å—Ç–∏ –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏
        for insight in insights['optimization_opportunities']:
            insights_data.append({'–ö–∞—Ç–µ–≥–æ—Ä–∏—è': '–û–ø—Ç–∏–º–∏–∑–∞—Ü–∏—è', '–ò–Ω—Å–∞–π—Ç': insight})
        
        # –ü—Ä–µ–¥—É–ø—Ä–µ–∂–¥–µ–Ω–∏—è
        for insight in insights['risk_alerts']:
            insights_data.append({'–ö–∞—Ç–µ–≥–æ—Ä–∏—è': '–†–∏—Å–∫–∏', '–ò–Ω—Å–∞–π—Ç': insight})
        
        # –§–∞–∫—Ç–æ—Ä—ã —É—Å–ø–µ—Ö–∞
        for insight in insights['success_factors']:
            insights_data.append({'–ö–∞—Ç–µ–≥–æ—Ä–∏—è': '–£—Å–ø–µ—Ö', '–ò–Ω—Å–∞–π—Ç': insight})
        
        insights_df = pd.DataFrame(insights_data)
        insights_df.to_excel(writer, sheet_name=ws_name, index=False, startrow=2)
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
        worksheet = writer.sheets[ws_name]
        worksheet['A1'] = '–ê–í–¢–û–ú–ê–¢–ò–ß–ï–°–ö–ò–ï –ò–ù–°–ê–ô–¢–´ –ò –†–ï–ö–û–ú–ï–ù–î–ê–¶–ò–ò'
        worksheet['A1'].font = Font(size=14, bold=True)
        worksheet.merge_cells('A1:B1')
        
        self._style_excel_table(worksheet, start_row=3, end_row=3+len(insights_df), 
                               start_col=1, end_col=2)
    
    def _style_excel_table(self, worksheet, start_row, end_row, start_col, end_col):
        """–ü—Ä–∏–º–µ–Ω–µ–Ω–∏–µ —Å—Ç–∏–ª–µ–π –∫ —Ç–∞–±–ª–∏—Ü–µ –≤ Excel."""
        # –°—Ç–∏–ª—å –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # –ì—Ä–∞–Ω–∏—Ü—ã —Ç–∞–±–ª–∏—Ü—ã
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # –ê–≤—Ç–æ—à–∏—Ä–∏–Ω–∞ –∫–æ–ª–æ–Ω–æ–∫
        for col in range(start_col, end_col + 1):
            column_letter = openpyxl.utils.get_column_letter(col)
            worksheet.column_dimensions[column_letter].auto_size = True
    
    def _add_chart_to_sheet(self, worksheet, data_df, position, chart_type='bar'):
        """–î–æ–±–∞–≤–ª–µ–Ω–∏–µ –¥–∏–∞–≥—Ä–∞–º–º—ã –≤ Excel –ª–∏—Å—Ç."""
        if chart_type == 'bar' and len(data_df) > 0:
            chart = BarChart()
            chart.title = "–†–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ –ø–æ –∫–∞–Ω–∞–ª–∞–º"
            chart.y_axis.title = "–ó–Ω–∞—á–µ–Ω–∏–µ"
            chart.x_axis.title = "–ö–∞–Ω–∞–ª—ã"
            
            # –î–∞–Ω–Ω—ã–µ –¥–ª—è –¥–∏–∞–≥—Ä–∞–º–º—ã
            data = Reference(worksheet, min_col=3, min_row=3, max_row=3+len(data_df), max_col=3)
            categories = Reference(worksheet, min_col=1, min_row=4, max_row=3+len(data_df))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            worksheet.add_chart(chart, position)
    
    def _interpret_metric(self, metric_name, value):
        """–ò–Ω—Ç–µ—Ä–ø—Ä–µ—Ç–∞—Ü–∏—è –º–µ—Ç—Ä–∏–∫–∏ –¥–ª—è –±–∏–∑–Ω–µ—Å–∞."""
        if '–ö–∞—á–µ—Å—Ç–≤–æ' in metric_name and isinstance(value, (int, float)):
            if value >= 0.8:
                return "–û—Ç–ª–∏—á–Ω–æ–µ –∫–∞—á–µ—Å—Ç–≤–æ"
            elif value >= 0.7:
                return "–•–æ—Ä–æ—à–µ–µ –∫–∞—á–µ—Å—Ç–≤–æ"
            elif value >= 0.5:
                return "–£–¥–æ–≤–ª–µ—Ç–≤–æ—Ä–∏—Ç–µ–ª—å–Ω–æ–µ"
            else:
                return "–¢—Ä–µ–±—É–µ—Ç —É–ª—É—á—à–µ–Ω–∏—è"
        
        elif '–¢–æ—á–Ω–æ—Å—Ç—å' in metric_name and isinstance(value, (int, float)):
            if value >= 85:
                return "–í—ã—Å–æ–∫–∞—è —Ç–æ—á–Ω–æ—Å—Ç—å"
            elif value >= 75:
                return "–•–æ—Ä–æ—à–∞—è —Ç–æ—á–Ω–æ—Å—Ç—å"
            elif value >= 60:
                return "–ü—Ä–∏–µ–º–ª–µ–º–∞—è —Ç–æ—á–Ω–æ—Å—Ç—å"
            else:
                return "–ù–∏–∑–∫–∞—è —Ç–æ—á–Ω–æ—Å—Ç—å"
        
        return "–ê–Ω–∞–ª–∏–∑–∏—Ä–æ–≤–∞—Ç—å –∏–Ω–¥–∏–≤–∏–¥—É–∞–ª—å–Ω–æ"
    
    def export_to_pdf(self, model_results, filename=None):
        """
        –≠–∫—Å–ø–æ—Ä—Ç —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ MMM –≤ PDF —Ñ–∞–π–ª.
        
        Parameters:
        -----------
        model_results : dict
            –°–ª–æ–≤–∞—Ä—å —Å —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏ –º–æ–¥–µ–ª–∏
        filename : str
            –ò–º—è —Ñ–∞–π–ª–∞ (–æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–æ)
        """
        if not PDF_AVAILABLE:
            raise ImportError("–î–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞ –≤ PDF —É—Å—Ç–∞–Ω–æ–≤–∏—Ç–µ: pip install reportlab")
        
        if filename is None:
            filename = f"MMM_Report_{self.timestamp}.pdf"
        
        # –ü—Ä–æ–≤–µ—Ä–∫–∞ –≤—Ö–æ–¥–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö
        if model_results is None:
            model_results = {}
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # –°—Ç–∏–ª–∏
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#4472C4')
        )
        
        # –ó–∞–≥–æ–ª–æ–≤–æ–∫ –æ—Ç—á–µ—Ç–∞
        story.append(Paragraph("–û–¢–ß–ï–¢ –ü–û MARKETING MIX MODEL", title_style))
        story.append(Paragraph(f"–î–∞—Ç–∞ —Å–æ–∑–¥–∞–Ω–∏—è: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # –ò—Å–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ–µ —Ä–µ–∑—é–º–µ
        story.append(Paragraph("–ò–°–ü–û–õ–ù–ò–¢–ï–õ–¨–ù–û–ï –†–ï–ó–Æ–ú–ï", heading_style))
        
        summary_text = f"""
        <b>–ö–∞—á–µ—Å—Ç–≤–æ –º–æ–¥–µ–ª–∏:</b> R¬≤ = {self._safe_get(model_results, 'r2_score', '–ù/–î')}<br/>
        <b>–¢–æ—á–Ω–æ—Å—Ç—å –ø—Ä–æ–≥–Ω–æ–∑–∞:</b> {self._safe_get(model_results, 'accuracy', 0):.1f}%<br/>
        <b>–°—Ç–∞—Ç—É—Å –º–æ–¥–µ–ª–∏:</b> {self._safe_get(model_results, 'model_status', '–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–æ')}<br/>
        <b>–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –∫–∞–Ω–∞–ª–æ–≤:</b> {len(self._safe_get(model_results, 'media_channels', []))}<br/>
        <b>–°—Ä–µ–¥–Ω–∏–π ROAS:</b> {self._safe_get(model_results, 'avg_roas', 0):.2f}<br/>
        """
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # –ö–ª—é—á–µ–≤—ã–µ –∏–Ω—Å–∞–π—Ç—ã
        insights = self.create_automated_insights(model_results)
        
        story.append(Paragraph("–ö–õ–Æ–ß–ï–í–´–ï –ò–ù–°–ê–ô–¢–´", heading_style))
        
        for insight in insights['performance_insights'][:3]:  # –¢–æ–ø 3 –∏–Ω—Å–∞–π—Ç–∞
            story.append(Paragraph(f"‚Ä¢ {insight}", styles['Normal']))
        
        story.append(Spacer(1, 15))
        
        # –î–µ–∫–æ–º–ø–æ–∑–∏—Ü–∏—è –ø—Ä–æ–¥–∞–∂
        contributions = self._safe_get(model_results, 'contributions', {})
        if contributions:
            story.append(Paragraph("–î–ï–ö–û–ú–ü–û–ó–ò–¶–ò–Ø –ü–†–û–î–ê–ñ", heading_style))
            
            contrib_data = [['–ö–∞–Ω–∞–ª', '–í–∫–ª–∞–¥', '–î–æ–ª—è (%)']]
            total = sum(contributions.values())
            
            for channel, value in contributions.items():
                contrib_data.append([
                    str(channel).replace('_spend', '').replace('_', ' ').title(),
                    f"{int(float(value)) if value is not None else 0:,}",
                    f"{float(value)/total*100:.1f}%" if total > 0 and value is not None else "0%"
                ])
            
            table = Table(contrib_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
        
        # ROAS –∞–Ω–∞–ª–∏–∑
        roas_data = self._safe_get(model_results, 'roas_data')
        if roas_data is not None and not self._safe_access_df(roas_data, 'empty'):
            story.append(Paragraph("–ê–ù–ê–õ–ò–ó ROAS", heading_style))
            
            if isinstance(roas_data, pd.DataFrame) and not roas_data.empty:
                roas_table_data = [['–ö–∞–Ω–∞–ª', 'ROAS', '–û—Ü–µ–Ω–∫–∞']]
                
                for _, row in roas_data.iterrows():
                    roas_val = float(row.get('ROAS', 0))
                    assessment = '–û—Ç–ª–∏—á–Ω–æ' if roas_val >= 3.0 else '–•–æ—Ä–æ—à–æ' if roas_val >= 2.0 else '–ü—Ä–∏–µ–º–ª–µ–º–æ' if roas_val >= 1.5 else '–ü–ª–æ—Ö–æ'
                    
                    roas_table_data.append([
                        str(row.get('Channel', '–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–æ')),
                        f"{roas_val:.2f}",
                        assessment
                    ])
                
                roas_table = Table(roas_table_data)
                roas_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(roas_table)
                story.append(Spacer(1, 20))
        
        # –†–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏
        story.append(Paragraph("–†–ï–ö–û–ú–ï–ù–î–ê–¶–ò–ò", heading_style))
        
        recommendations = self._generate_recommendations(model_results)
        for rec in recommendations:
            story.append(Paragraph(f"‚Ä¢ {rec}", styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # –†–∏—Å–∫–∏ –∏ –≤–æ–∑–º–æ–∂–Ω–æ—Å—Ç–∏
        story.append(Paragraph("–†–ò–°–ö–ò –ò –í–û–ó–ú–û–ñ–ù–û–°–¢–ò", heading_style))
        
        for risk in insights['risk_alerts']:
            story.append(Paragraph(f"‚ö†Ô∏è {risk}", styles['Normal']))
        
        for opp in insights['optimization_opportunities']:
            story.append(Paragraph(f"üìà {opp}", styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # –§—É—Ç–µ—Ä
        footer_text = "–°–æ–∑–¥–∞–Ω–æ —Å –ø–æ–º–æ—â—å—é Marketing Mix Model Analytics Platform v2.1"
        story.append(Paragraph(footer_text, styles['Normal']))
        
        # –°–±–æ—Ä–∫–∞ PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer.getvalue(), filename
    
    def _generate_recommendations(self, model_results):
        """–ì–µ–Ω–µ—Ä–∞—Ü–∏—è —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–π –Ω–∞ –æ—Å–Ω–æ–≤–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ –º–æ–¥–µ–ª–∏."""
        recommendations = []
        
        # –†–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏ –ø–æ –∫–∞—á–µ—Å—Ç–≤—É –º–æ–¥–µ–ª–∏
        r2_score = self._safe_get(model_results, 'r2_score', 0)
        if r2_score >= 0.8:
            recommendations.append("–ú–æ–¥–µ–ª—å –¥–µ–º–æ–Ω—Å—Ç—Ä–∏—Ä—É–µ—Ç –≤—ã—Å–æ–∫–æ–µ –∫–∞—á–µ—Å—Ç–≤–æ. –†–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏ –º–æ–∂–Ω–æ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å –¥–ª—è –ø–ª–∞–Ω–∏—Ä–æ–≤–∞–Ω–∏—è –±—é–¥–∂–µ—Ç–∞.")
        elif r2_score >= 0.5:
            recommendations.append("–ú–æ–¥–µ–ª—å –ø–æ–∫–∞–∑—ã–≤–∞–µ—Ç –ø—Ä–∏–µ–º–ª–µ–º–æ–µ –∫–∞—á–µ—Å—Ç–≤–æ. –ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏ —Å –æ—Å—Ç–æ—Ä–æ–∂–Ω–æ—Å—Ç—å—é.")
        else:
            recommendations.append("–ö–∞—á–µ—Å—Ç–≤–æ –º–æ–¥–µ–ª–∏ –Ω–∏–∑–∫–æ–µ. –¢—Ä–µ–±—É–µ—Ç—Å—è —É–ª—É—á—à–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã—Ö –∏–ª–∏ –ø–∞—Ä–∞–º–µ—Ç—Ä–æ–≤ –º–æ–¥–µ–ª–∏.")
        
        # –†–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏ –ø–æ ROAS
        avg_roas = self._safe_get(model_results, 'avg_roas', 0)
        if avg_roas >= 3.0:
            recommendations.append("–°—Ä–µ–¥–Ω–∏–π ROAS –≤—ã—Å–æ–∫–∏–π. –†–∞—Å—Å–º–æ—Ç—Ä–∏—Ç–µ —É–≤–µ–ª–∏—á–µ–Ω–∏–µ –æ–±—â–µ–≥–æ –±—é–¥–∂–µ—Ç–∞.")
        elif avg_roas >= 2.0:
            recommendations.append("ROAS –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –≤ –ø—Ä–∏–µ–º–ª–µ–º–æ–º –¥–∏–∞–ø–∞–∑–æ–Ω–µ. –§–æ–∫—É—Å–∏—Ä—É–π—Ç–µ—Å—å –Ω–∞ –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏ —Ä–∞—Å–ø—Ä–µ–¥–µ–ª–µ–Ω–∏—è.")
        else:
            recommendations.append("ROAS –Ω–∏–∂–µ –æ–∂–∏–¥–∞–µ–º–æ–≥–æ. –¢—Ä–µ–±—É–µ—Ç—Å—è –ø–µ—Ä–µ—Å–º–æ—Ç—Ä —Å—Ç—Ä–∞—Ç–µ–≥–∏–∏ –∏–ª–∏ –∫–∞–Ω–∞–ª–æ–≤.")
        
        # –†–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏ –ø–æ –¥–µ–∫–æ–º–ø–æ–∑–∏—Ü–∏–∏
        base_pct = self._safe_get(model_results, 'base_contribution_pct', 0)
        if base_pct > 70:
            recommendations.append("–í—ã—Å–æ–∫–∞—è –¥–æ–ª—è –±–∞–∑–æ–≤–æ–π –ª–∏–Ω–∏–∏. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ —ç—Ñ—Ñ–µ–∫—Ç–∏–≤–Ω–æ—Å—Ç—å –º–µ–¥–∏–∞-–∫–∞–Ω–∞–ª–æ–≤.")
        elif base_pct < 30:
            recommendations.append("–ù–∏–∑–∫–∞—è –±–∞–∑–æ–≤–∞—è –ª–∏–Ω–∏—è. –í–æ–∑–º–æ–∂–Ω–æ –ø–µ—Ä–µ–æ—Ü–µ–Ω–∏–≤–∞–µ—Ç—Å—è –≤–ª–∏—è–Ω–∏–µ –º–µ–¥–∏–∞.")
        else:
            recommendations.append("–ó–¥–æ—Ä–æ–≤–æ–µ —Å–æ–æ—Ç–Ω–æ—à–µ–Ω–∏–µ –º–µ–∂–¥—É –±–∞–∑–æ–≤–æ–π –ª–∏–Ω–∏–µ–π –∏ –º–µ–¥–∏–∞-—ç—Ñ—Ñ–µ–∫—Ç–∞–º–∏.")
        
        return recommendations
    
    def export_quick_summary(self, model_results, format='excel'):
        """
        –ë—ã—Å—Ç—Ä—ã–π —ç–∫—Å–ø–æ—Ä—Ç –∫—Ä–∞—Ç–∫–æ–π —Å–≤–æ–¥–∫–∏ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤.
        
        Parameters:
        -----------
        model_results : dict
            –°–ª–æ–≤–∞—Ä—å —Å —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏ –º–æ–¥–µ–ª–∏
        format : str
            –§–æ—Ä–º–∞—Ç —ç–∫—Å–ø–æ—Ä—Ç–∞ ('excel' –∏–ª–∏ 'pdf')
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # –ü—Ä–æ–≤–µ—Ä–∫–∞ –≤—Ö–æ–¥–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö
        if model_results is None:
            model_results = {}
        
        if format == 'excel':
            filename = f"MMM_Summary_{timestamp}.xlsx"
            buffer = BytesIO()
            
            # –°–æ–∑–¥–∞–Ω–∏–µ –ø—Ä–æ—Å—Ç–æ–≥–æ Excel —Å –æ—Å–Ω–æ–≤–Ω—ã–º–∏ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞–º–∏
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # –û—Å–Ω–æ–≤–Ω—ã–µ –º–µ—Ç—Ä–∏–∫–∏
                summary_data = {
                    '–ú–µ—Ç—Ä–∏–∫–∞': [
                        '–ö–∞—á–µ—Å—Ç–≤–æ –º–æ–¥–µ–ª–∏ (R¬≤)',
                        '–¢–æ—á–Ω–æ—Å—Ç—å –ø—Ä–æ–≥–Ω–æ–∑–∞ (%)',
                        '–°—Ä–µ–¥–Ω–∏–π ROAS',
                        '–û–±—â–∏–π –±—é–¥–∂–µ—Ç',
                        '–°—Ç–∞—Ç—É—Å –º–æ–¥–µ–ª–∏'
                    ],
                    '–ó–Ω–∞—á–µ–Ω–∏–µ': [
                        f"{self._safe_get(model_results, 'r2_score', 0):.3f}",
                        f"{self._safe_get(model_results, 'accuracy', 0):.1f}%",
                        f"{self._safe_get(model_results, 'avg_roas', 0):.2f}",
                        f"{self._safe_get(model_results, 'total_budget', 0):,.0f} —Ä—É–±",
                        self._safe_get(model_results, 'model_status', '–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–æ')
                    ]
                }
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='–°–≤–æ–¥–∫–∞', index=False)
                
                # ROAS –¥–∞–Ω–Ω—ã–µ –µ—Å–ª–∏ –¥–æ—Å—Ç—É–ø–Ω—ã
                roas_data = self._safe_get(model_results, 'roas_data')
                if roas_data is not None and not self._safe_access_df(roas_data, 'empty'):
                    roas_data.to_excel(writer, sheet_name='ROAS', index=False)
            
            buffer.seek(0)
            return buffer.getvalue(), filename
        
        elif format == 'pdf':
            filename = f"MMM_Summary_{timestamp}.pdf"
            buffer = BytesIO()
            
            # –°–æ–∑–¥–∞–Ω–∏–µ –ø—Ä–æ—Å—Ç–æ–≥–æ PDF
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # –ó–∞–≥–æ–ª–æ–≤–æ–∫
            story.append(Paragraph("–ö–†–ê–¢–ö–ê–Ø –°–í–û–î–ö–ê MMM", styles['Title']))
            story.append(Spacer(1, 20))
            
            # –û—Å–Ω–æ–≤–Ω—ã–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã
            summary_text = f"""
            <b>–ö–∞—á–µ—Å—Ç–≤–æ –º–æ–¥–µ–ª–∏:</b> R¬≤ = {self._safe_get(model_results, 'r2_score', 0):.3f}<br/>
            <b>–¢–æ—á–Ω–æ—Å—Ç—å:</b> {self._safe_get(model_results, 'accuracy', 0):.1f}%<br/>
            <b>ROAS:</b> {self._safe_get(model_results, 'avg_roas', 0):.2f}<br/>
            <b>–°—Ç–∞—Ç—É—Å:</b> {self._safe_get(model_results, 'model_status', '–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–æ')}<br/>
            """
            story.append(Paragraph(summary_text, styles['Normal']))
            
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue(), filename
    
    def create_automated_insights(self, model_results):
        """
        –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∞—è –≥–µ–Ω–µ—Ä–∞—Ü–∏—è –∏–Ω—Å–∞–π—Ç–æ–≤ –Ω–∞ –æ—Å–Ω–æ–≤–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ –º–æ–¥–µ–ª–∏.
        
        Returns:
        --------
        dict : –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –∏–Ω—Å–∞–π—Ç—ã
        """
        insights = {
            'performance_insights': [],
            'optimization_opportunities': [],
            'risk_alerts': [],
            'success_factors': []
        }
        
        # –ë–µ–∑–æ–ø–∞—Å–Ω–æ–µ –∏–∑–≤–ª–µ—á–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã—Ö
        r2_score = self._safe_get(model_results, 'r2_score', 0)
        avg_roas = self._safe_get(model_results, 'avg_roas', 0)
        contributions = self._safe_get(model_results, 'contributions', {})
        
        # –ò–Ω—Å–∞–π—Ç—ã –ø–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å–Ω–æ—Å—Ç–∏
        if r2_score >= 0.8:
            insights['performance_insights'].append("–ú–æ–¥–µ–ª—å –¥–µ–º–æ–Ω—Å—Ç—Ä–∏—Ä—É–µ—Ç –≤—ã—Å–æ–∫—É—é —Ç–æ—á–Ω–æ—Å—Ç—å –ø—Ä–µ–¥—Å–∫–∞–∑–∞–Ω–∏–π")
        
        if avg_roas >= 3.0:
            insights['performance_insights'].append("–ú–∞—Ä–∫–µ—Ç–∏–Ω–≥–æ–≤—ã–µ –∏–Ω–≤–µ—Å—Ç–∏—Ü–∏–∏ –ø–æ–∫–∞–∑—ã–≤–∞—é—Ç –æ—Ç–ª–∏—á–Ω—É—é –æ–∫—É–ø–∞–µ–º–æ—Å—Ç—å")
        
        # –í–æ–∑–º–æ–∂–Ω–æ—Å—Ç–∏ –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏
        roas_data = self._safe_get(model_results, 'roas_data')
        if roas_data is not None and not self._safe_access_df(roas_data, 'empty'):
            try:
                best_channel = roas_data.loc[roas_data['ROAS'].idxmax()]
                worst_channel = roas_data.loc[roas_data['ROAS'].idxmin()]
                
                insights['optimization_opportunities'].append(
                    f"–õ—É—á—à–∏–π –∫–∞–Ω–∞–ª ({best_channel['Channel']}) –ø–æ–∫–∞–∑—ã–≤–∞–µ—Ç ROAS {best_channel['ROAS']:.2f} - —Ä–∞—Å—Å–º–æ—Ç—Ä–µ—Ç—å —É–≤–µ–ª–∏—á–µ–Ω–∏–µ –±—é–¥–∂–µ—Ç–∞"
                )
                
                if worst_channel['ROAS'] < 1.5:
                    insights['optimization_opportunities'].append(
                        f"–ö–∞–Ω–∞–ª {worst_channel['Channel']} –ø–æ–∫–∞–∑—ã–≤–∞–µ—Ç –Ω–∏–∑–∫–∏–π ROAS {worst_channel['ROAS']:.2f} - —Ç—Ä–µ–±—É–µ—Ç –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏"
                    )
            except (KeyError, IndexError, TypeError):
                insights['optimization_opportunities'].append("–†–µ–∫–æ–º–µ–Ω–¥—É–µ—Ç—Å—è –¥–µ—Ç–∞–ª—å–Ω—ã–π –∞–Ω–∞–ª–∏–∑ —ç—Ñ—Ñ–µ–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –∫–∞–Ω–∞–ª–æ–≤")
        
        # –ü—Ä–µ–¥—É–ø—Ä–µ–∂–¥–µ–Ω–∏—è –æ —Ä–∏—Å–∫–∞—Ö
        if r2_score < 0.6:
            insights['risk_alerts'].append("–ù–∏–∑–∫–æ–µ –∫–∞—á–µ—Å—Ç–≤–æ –º–æ–¥–µ–ª–∏ - —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏ –º–æ–≥—É—Ç –±—ã—Ç—å –Ω–µ—Ç–æ—á–Ω—ã–º–∏")
        
        base_share = 0
        if contributions:
            total_contribution = sum(contributions.values())
            if total_contribution > 0:
                base_share = contributions.get('Base', 0) / total_contribution
                
                if base_share > 0.8:
                    insights['risk_alerts'].append("–í—ã—Å–æ–∫–∞—è –¥–æ–ª—è –æ—Ä–≥–∞–Ω–∏—á–µ—Å–∫–∏—Ö –ø—Ä–æ–¥–∞–∂ - –≤–æ–∑–º–æ–∂–Ω–∞ –Ω–µ–¥–æ–æ—Ü–µ–Ω–∫–∞ –º–µ–¥–∏–∞-—ç—Ñ—Ñ–µ–∫—Ç–æ–≤")
                elif base_share < 0.2:
                    insights['risk_alerts'].append("–ù–∏–∑–∫–∞—è –¥–æ–ª—è –æ—Ä–≥–∞–Ω–∏—á–µ—Å–∫–∏—Ö –ø—Ä–æ–¥–∞–∂ - –≤–æ–∑–º–æ–∂–Ω–∞ –ø–µ—Ä–µ–æ—Ü–µ–Ω–∫–∞ –º–µ–¥–∏–∞-—ç—Ñ—Ñ–µ–∫—Ç–æ–≤")
        
        # –§–∞–∫—Ç–æ—Ä—ã —É—Å–ø–µ—Ö–∞
        if avg_roas >= 2.5:
            insights['success_factors'].append("–≠—Ñ—Ñ–µ–∫—Ç–∏–≤–Ω–∞—è –º–µ–¥–∏–∞-—Å—Ç—Ä–∞—Ç–µ–≥–∏—è —Å –≤—ã—Å–æ–∫–∏–º ROAS")
        
        if 0.3 <= base_share <= 0.6:
            insights['success_factors'].append("–ó–¥–æ—Ä–æ–≤—ã–π –±–∞–ª–∞–Ω—Å –º–µ–∂–¥—É –æ—Ä–≥–∞–Ω–∏—á–µ—Å–∫–∏–º–∏ –∏ –º–µ–¥–∏–∞-–ø—Ä–æ–¥–∞–∂–∞–º–∏")
        
        return insights
    
    def create_export_data(self, model, data, contributions, roas_data, metrics, optimization_results=None, scenarios=None):
        """
        –ü–æ–¥–≥–æ—Ç–æ–≤–∫–∞ –¥–∞–Ω–Ω—ã—Ö –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞.
        
        Parameters:
        -----------
        model : MarketingMixModel
            –û–±—É—á–µ–Ω–Ω–∞—è –º–æ–¥–µ–ª—å
        data : pd.DataFrame
            –ò—Å—Ö–æ–¥–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ
        contributions : dict
            –í–∫–ª–∞–¥—ã –∫–∞–Ω–∞–ª–æ–≤
        roas_data : pd.DataFrame
            –î–∞–Ω–Ω—ã–µ ROAS
        metrics : dict
            –ú–µ—Ç—Ä–∏–∫–∏ –∫–∞—á–µ—Å—Ç–≤–∞ –º–æ–¥–µ–ª–∏
        optimization_results : dict
            –†–µ–∑—É–ª—å—Ç–∞—Ç—ã –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏ (–æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–æ)
        scenarios : dict
            –†–µ–∑—É–ª—å—Ç–∞—Ç—ã —Å—Ü–µ–Ω–∞—Ä–∏–µ–≤ (–æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–æ)
        """
        
        # –ë–∞–∑–æ–≤—ã–µ —Ä–∞—Å—á–µ—Ç—ã —Å –±–µ–∑–æ–ø–∞—Å–Ω–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–æ–π
        total_contribution = sum(contributions.values()) if contributions else 1
        base_contribution = contributions.get('Base', 0) if contributions else 0
        base_pct = (base_contribution / total_contribution * 100) if total_contribution > 0 else 0
        
        avg_roas = 0
        if roas_data is not None and not self._safe_access_df(roas_data, 'empty'):
            if 'ROAS' in roas_data.columns:
                avg_roas = roas_data['ROAS'].mean()
        
        # –û–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ —Å—Ç–∞—Ç—É—Å–∞ –º–æ–¥–µ–ª–∏
        r2_score = metrics.get('–ö–∞—á–µ—Å—Ç–≤–æ –ø—Ä–æ–≥–Ω–æ–∑–∞', 0) if metrics else 0
        accuracy = metrics.get('–¢–æ—á–Ω–æ—Å—Ç—å –º–æ–¥–µ–ª–∏ (%)', 0) if metrics else 0
        
        if r2_score >= 0.8 and accuracy >= 85:
            model_status = "–û—Ç–ª–∏—á–Ω–æ - –≥–æ—Ç–æ–≤–∞ –∫ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—é"
        elif r2_score >= 0.7 and accuracy >= 75:
            model_status = "–•–æ—Ä–æ—à–æ - –ø–æ–¥—Ö–æ–¥–∏—Ç –¥–ª—è –ø–ª–∞–Ω–∏—Ä–æ–≤–∞–Ω–∏—è"
        elif r2_score >= 0.5 and accuracy >= 60:
            model_status = "–£–¥–æ–≤–ª–µ—Ç–≤–æ—Ä–∏—Ç–µ–ª—å–Ω–æ - –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å –æ—Å—Ç–æ—Ä–æ–∂–Ω–æ"
        else:
            model_status = "–ü–ª–æ—Ö–æ - —Ç—Ä–µ–±—É–µ—Ç —É–ª—É—á—à–µ–Ω–∏—è"
        
        # –ë–µ–∑–æ–ø–∞—Å–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –º–µ–¥–∏–∞-–∫–∞–Ω–∞–ª–æ–≤
        media_channels = []
        if hasattr(model, 'media_channels') and model.media_channels:
            media_channels = model.media_channels
        
        # –ë–µ–∑–æ–ø–∞—Å–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –ø–µ—Ä–∏–æ–¥–∞ –∞–Ω–∞–ª–∏–∑–∞
        analysis_period = '–ù/–î'
        if data is not None and not self._safe_access_df(data, 'empty') and 'date' in data.columns:
            try:
                analysis_period = f"{data['date'].min().strftime('%Y-%m-%d')} - {data['date'].max().strftime('%Y-%m-%d')}"
            except:
                analysis_period = '–ù/–î'
        
        # –ë–µ–∑–æ–ø–∞—Å–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –æ–±—â–µ–≥–æ –±—é–¥–∂–µ—Ç–∞
        total_budget = 0
        if data is not None and not self._safe_access_df(data, 'empty'):
            for ch in media_channels:
                if ch in data.columns:
                    try:
                        total_budget += data[ch].sum()
                    except:
                        pass
        
        export_data = {
            'r2_score': r2_score,
            'accuracy': accuracy,
            'model_status': model_status,
            'media_channels': media_channels,
            'analysis_period': analysis_period,
            'total_budget': total_budget,
            'avg_roas': avg_roas,
            'base_contribution_pct': base_pct,
            'contributions': contributions or {},
            'roas_data': roas_data,
            'model_metrics': metrics or {},
            'raw_data': data
        }
        
        if optimization_results:
            export_data['optimization_results'] = optimization_results
        
        if scenarios:
            export_data['scenarios'] = scenarios
        
        return export_data
